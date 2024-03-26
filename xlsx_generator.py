import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from api_requests import fetch_data

# Debugging here to inspect all_words data
# import pdb
# pdb.set_trace()

def fetch_data_from_api():
    # Fetch lessons
    lessons = fetch_data({'course_id': 5}, 'lesson')

    all_words = []
    for lesson in lessons:
        words = fetch_data({'lesson_id': lesson['id']}, 'topic')
        for word in words:
            word['lesson_title'] = lesson['title']  # Add lesson title to each word
        all_words.extend(words)
    
    # Return both lessons and all_words
    return lessons, all_words

def filter_and_organize_data(all_words):
    # Create a copy to avoid SettingWithCopyWarning
    df_words = pd.DataFrame(all_words).copy()
    
    # Select only the necessary columns and add the new columns you requested
    df_words_filtered = df_words[['lesson_title', 'content', 'trans', 'sentence1', 'vi_sentence1', 'phonetic', 'position', 'audio', 'picture']].copy()
    
    # Rename columns for easier understanding
    df_words_filtered.rename(columns={
        'lesson_title': 'Topic',
        'content': 'Word',
        'trans': 'Meaning',
        'sentence1': 'Example Sentence (EN)',
        'vi_sentence1': 'Example Sentence (VI)',
        'phonetic': 'Phonetic',
        'position': 'Position',
        'audio': 'Audio',
        'picture': 'Picture'
    }, inplace=True)
    
    return df_words_filtered

def generate_excel_with_merged_cells(lessons, df_words_filtered):
    # Tạo dữ liệu cho Excel
    excel_data = []

    grouped = df_words_filtered.groupby('Topic')

    for lesson in lessons:
        lesson_title = lesson['title']
        lesson_description = lesson.get('description', '')  # Lấy mô tả của bài học
        lesson_image = lesson.get('image', '')
        
        if lesson_title in grouped.groups:
            lesson_words = grouped.get_group(lesson_title)
            for _, row in lesson_words.iterrows():
                excel_data.append({
                    'Lesson': lesson_title,
                    'Description': lesson_description,
                    'Image': lesson_image,
                    'Word': row['Word'],
                    'Phonetic': row['Phonetic'],
                    'Position': row['Position'],
                    'Meaning': row['Meaning'],
                    'Example Sentence (EN)': row['Example Sentence (EN)'],
                    'Example Sentence (VI)': row['Example Sentence (VI)'],
                    'Audio': row['Audio'],
                    'Picture': row['Picture']
                })

    # Convert to DataFrame
    df_excel = pd.DataFrame(excel_data)

    # Xuất dữ liệu ra file Excel
    excel_file = "mindmap_data.xlsx"
    df_excel.to_excel(excel_file, index=False)

    # Load lại file Excel bằng openpyxl để merge cells
    wb = load_workbook(excel_file)
    ws = wb.active  # Lấy worksheet đầu tiên
    
    # Merge các ô trong cột 'Lesson' khi có giá trị lặp lại
    merge_cells(ws, 'A')  # 'A' là cột Lesson
    merge_cells(ws, 'B')  # 'B' là cột Description
    merge_cells(ws, 'C')  # 'C' là cột Image

    # Lưu lại file Excel
    wb.save(excel_file)
    print(f"Excel file '{excel_file}' has been generated with merged cells for 'Lesson' and 'Description'.")

def merge_cells(ws: Worksheet, col_letter: str):
    """
    Hàm để merge các ô trong một cột nếu có giá trị lặp lại.
    col_letter: Chữ cái của cột (ví dụ 'A' cho cột Lesson)
    """
    current_value = None
    start_row = None
    
    for row in range(2, ws.max_row + 1):  # Bắt đầu từ hàng 2 để bỏ qua tiêu đề
        cell_value = ws[f"{col_letter}{row}"].value
        
        if cell_value != current_value:
            if current_value is not None and start_row is not None and start_row != row - 1:
                # Merge các ô từ hàng bắt đầu đến hàng hiện tại - 1
                ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{row-1}")
            # Cập nhật giá trị mới
            current_value = cell_value
            start_row = row
    # Merge các ô cuối cùng nếu cần
    if start_row is not None and start_row != ws.max_row:
        ws.merge_cells(f"{col_letter}{start_row}:{col_letter}{ws.max_row}")

# Main logic
if __name__ == "__main__":
    # Step 1: Fetch data from API
    lessons, all_words = fetch_data_from_api()

    # Step 2: Filter and organize data
    df_words_filtered = filter_and_organize_data(all_words)

    # Step 3: Generate mind map
    generate_excel_with_merged_cells(lessons, df_words_filtered)
